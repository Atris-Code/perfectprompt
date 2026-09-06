"""
Excel export for CFO Project Finance simulations.

Generates a multi-sheet .xlsx report with inputs, revenue stack, cash flow
waterfall, sensitivities and executive KPIs. All figures are computed values
from the simulation engine (no broken/circular formulas).
"""

import io
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.cfo.specs import FinancialParametersSpec, IndustrialProcessSpec, SimulationResult


def _style_header(ws, row: int = 1) -> None:
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin


def _autosize(ws, widths: Dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def build_excel_report(
    process_spec: IndustrialProcessSpec,
    financial_spec: FinancialParametersSpec,
    result: SimulationResult,
    revenue_details: Dict[str, Any],
) -> bytes:
    """Build the .xlsx report and return its bytes."""
    wb = Workbook()
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ---- Sheet 1: INPUTS ----
    ws1 = wb.active
    ws1.title = "INPUTS"
    ws1.append(["PARÁMETRO", "VALOR", "UNIDAD"])
    _style_header(ws1)

    proc_fields = [
        ("Tipo de proceso", process_spec.process_type.value, "-"),
        ("Nombre del proyecto", process_spec.name, "-"),
        ("CAPEX fijo", round(process_spec.fixed_capex, 2), "€"),
        ("NWC", round(process_spec.nwc, 2), "€"),
        ("O&M fijo anual", round(process_spec.fixed_om_eur_year, 2), "€/año"),
        ("O&M variable", process_spec.variable_om_pct_revenue, "% ingresos"),
    ]
    fin_fields = list(financial_spec.to_dict().items()) if hasattr(financial_spec, "to_dict") else list(financial_spec.dict().items())

    for label, value, unit in proc_fields:
        ws1.append([label, value, unit])
    ws1.append([])
    ws1.append(["PARÁMETRO FINANCIERO", "VALOR", "UNIDAD"])
    _style_header(ws1, ws1.max_row)
    for key, value in fin_fields:
        ws1.append([key, value, "-"])

    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, max_col=3):
        for cell in row:
            cell.border = thin
    _autosize(ws1, {"A": 38, "B": 22, "C": 14})

    # ---- Sheet 2: REVENUE STACK ----
    ws2 = wb.create_sheet("REVENUE_STACK")
    ws2.append(["CONCEPTO", "VALOR", "UNIDAD"])
    _style_header(ws2)
    for key, value in revenue_details.items():
        ws2.append([key.replace("_", " ").title(), value, "€" if "revenue" in key or "cost" in key or "opex" in key else "-"])
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=3):
        for cell in row:
            cell.border = thin
    _autosize(ws2, {"A": 38, "B": 22, "C": 14})

    # ---- Sheet 3: CASHFLOW ----
    ws3 = wb.create_sheet("CASHFLOW")
    headers = ["AÑO", "INGRESO", "OPEX", "EBITDA", "DEPRECIACIÓN", "EBIT", "INTERESES",
               "EBT", "IMPUESTOS", "CFADS", "SERVICIO DEUDA", "PRINCIPAL", "DEUDA RESTANTE",
               "CASH SWEEP", "FCFE", "DSCR"]
    ws3.append(headers)
    _style_header(ws3)
    for p in result.annual_projections:
        ws3.append([
            p.year, p.revenue, p.opex, p.ebitda, p.depreciation, p.ebit, p.interest,
            p.ebt, p.tax, p.cfads, p.debt_service, p.principal, p.remaining_debt,
            p.cash_sweep, p.fcfe, p.dscr,
        ])
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = thin
            if cell.column > 1:
                cell.number_format = "#,##0.00"
    _autosize(ws3, {get_column_letter(i): 16 for i in range(1, len(headers) + 1)})

    # ---- Sheet 4: SENSIBILIDADES ----
    ws4 = wb.create_sheet("SENSIBILIDADES")
    ws4.append(["ESCENARIO", "EBITDA", "CFADS", "DSCR MIN", "DSCR PROM", "IRR EQUITY",
                "NPV EQUITY", "CASH SWEEP", "DEFAULT"])
    _style_header(ws4)
    for key, sc in result.sensitivities.items():
        ws4.append([
            sc.name, sc.ebitda, sc.cfads, sc.min_dscr, sc.avg_dscr,
            (round(sc.equity_irr * 100, 2) if sc.equity_irr is not None else None),
            sc.project_npv, sc.cash_sweep_triggered, sc.default_alert,
        ])
    for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row, max_col=9):
        for cell in row:
            cell.border = thin
            if cell.column in (2, 3, 4, 5, 7):
                cell.number_format = "#,##0.00"
            elif cell.column == 6:
                cell.number_format = "0.00"
    _autosize(ws4, {"A": 45, "B": 16, "C": 16, "D": 12, "E": 12, "F": 12, "G": 14, "H": 12, "I": 10})

    # ---- Sheet 5: RESUMEN ----
    ws5 = wb.create_sheet("RESUMEN")
    ws5.append(["KPI", "VALOR"])
    _style_header(ws5)
    s = result.summary
    kpis = [
        ("CAPEX total", round(s.total_capex, 2)),
        ("Equity invertido", round(s.equity_invested, 2)),
        ("Deuda senior", round(s.senior_debt_principal, 2)),
        ("Cuota anual deuda", round(s.senior_debt_annual_payment, 2)),
        ("EBITDA año 1", round(s.ebitda_base_year1, 2)),
        ("DSCR mínimo", round(s.min_dscr, 2)),
        ("DSCR promedio", round(s.avg_dscr, 2)),
        ("IRR Equity (%)", (round(s.equity_irr * 100, 2) if s.equity_irr is not None else None)),
        ("IRR Proyecto (%)", (round(s.project_irr * 100, 2) if s.project_irr is not None else None)),
        ("NPV Equity", round(s.npv_equity, 2) if s.npv_equity is not None else None),
        ("NPV Proyecto", round(s.project_npv, 2) if s.project_npv is not None else None),
        ("Payback dinámico (años)", s.dynamic_payback_years),
        ("Breaches de covenant", s.covenant_breaches_count),
        ("Cash sweep activado", s.cash_sweep_triggered),
        ("Default alert", s.default_alert),
    ]
    for label, value in kpis:
        ws5.append([label, value])
    for row in ws5.iter_rows(min_row=2, max_row=ws5.max_row, max_col=2):
        for cell in row:
            cell.border = thin
    _autosize(ws5, {"A": 32, "B": 22})

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
